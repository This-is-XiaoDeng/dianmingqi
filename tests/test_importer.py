"""importer 单元测试：txt 与 xlsx 解析。"""
import io

import pytest
from openpyxl import Workbook

from dianmingqi.importer import parse_txt, parse_xlsx_column


def test_parse_txt_basic(tmp_path):
    f = tmp_path / "list.txt"
    f.write_text("张三\n李四\n王五\n", encoding="utf-8")
    assert parse_txt(str(f)) == ["张三", "李四", "王五"]


def test_parse_txt_strip_and_empty(tmp_path):
    f = tmp_path / "list.txt"
    f.write_text("  张三  \n\n李四\n\n\n王五\n", encoding="utf-8")
    assert parse_txt(str(f)) == ["张三", "李四", "王五"]


def test_parse_txt_bom_and_dedup(tmp_path):
    f = tmp_path / "list.txt"
    f.write_text("\ufeff张三\n张三\n李四\n", encoding="utf-8")
    assert parse_txt(str(f)) == ["张三", "李四"]


def test_parse_xlsx_strip_header(tmp_path):
    wb = Workbook()
    ws = wb.active
    # 表头 + 数据
    ws.append(["姓名"])
    ws.append(["张三"])
    ws.append(["李四"])
    ws.append(["王五"])
    p = tmp_path / "list.xlsx"
    wb.save(p)
    # 默认第 1 列，去除表头
    assert parse_xlsx_column(str(p)) == ["张三", "李四", "王五"]


def test_parse_xlsx_select_column_and_empty(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["序号", "姓名", "备注"])
    ws.append([1, "张三", "甲"])
    ws.append([2, "", "乙"])
    ws.append([3, "李四", "丙"])
    ws.append([4, "张三", "丁"])  # 重复会去重
    p = tmp_path / "list.xlsx"
    wb.save(p)
    # 选第 2 列（姓名），去表头、去空、去重
    assert parse_xlsx_column(str(p), column=2) == ["张三", "李四"]


def test_parse_xlsx_all_header_only(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名"])
    p = tmp_path / "list.xlsx"
    wb.save(p)
    assert parse_xlsx_column(str(p)) == []


def test_parse_xlsx_has_header_checkbox(tmp_path):
    """has_header=False 时第一行（表头）也作为名字导入。"""
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名"])
    ws.append(["张三"])
    ws.append(["李四"])
    p = tmp_path / "list.xlsx"
    wb.save(p)
    assert parse_xlsx_column(str(p), has_header=True) == ["张三", "李四"]
    assert parse_xlsx_column(str(p), has_header=False) == ["姓名", "张三", "李四"]
