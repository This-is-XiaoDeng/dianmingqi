"""数据导入：从 txt / xlsx 读取名单。

- txt：一行一个，忽略空行，strip 首尾空白。
- xlsx：选择一个整列，去除表头（第一行）后取该列剩余单元格，
  忽略空单元格并去重、去空白。

对外只暴露 :func:`parse_file`,返回去重后的名单列表。
"""
from __future__ import annotations

import os
from typing import List, Optional

from openpyxl import load_workbook


def _clean(items: List[str]) -> List[str]:
    """去掉空白条目并保持顺序去重。"""
    seen = set()
    out: List[str] = []
    for raw in items:
        name = raw.strip()
        if not name:
            continue
        if name in seen:
            continue
        seen.add(name)
        out.append(name)
    return out


def parse_txt(path: str) -> List[str]:
    with open(path, "r", encoding="utf-8-sig") as f:
        lines = [line.strip() for line in f]
    return _clean(lines)


def parse_xlsx_column(path: str, sheet: Optional[int] = None, column: Optional[int] = None) -> List[str]:
    """从 xlsx 读取一整列并去除表头。

    :param sheet: 工作表序号（从 0 开始），默认第一个。
    :param column: 列序号（从 1 开始），默认第一列（A）。
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[sheet if sheet is not None else 0]
        col = column if column is not None else 1
        # 去除表头：跳过第一行
        values = []
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            cell = row[0]
            if cell is None:
                continue
            values.append(str(cell))
        # values[0] 是表头，去掉
        if values:
            values = values[1:]
        return _clean(values)
    finally:
        wb.close()


def parse_file(path: str, sheet: Optional[int] = None, column: Optional[int] = None) -> List[str]:
    """根据扩展名自动选择合适的解析器。"""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return parse_xlsx_column(path, sheet=sheet, column=column)
    # 其余按 txt 处理：一行一个
    return parse_txt(path)
